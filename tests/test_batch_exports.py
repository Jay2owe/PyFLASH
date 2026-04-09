from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from IF_analysis.batch import Batch
from IF_analysis.conditions import condition, conditionList
from IF_analysis.export import convert_summary_sheet_name
from IF_analysis.markers import Antibody


def _build_export_batch(tmp_path):
    cond = condition("A", "A", "#000000", "Genotype", explanation="<>")
    conditions = conditionList([cond])

    summary = pd.DataFrame(
        {
            "AnimalName": ["A1"],
            "Condition": ["A"],
            "GFAP_Count": [1.0],
            "GFAP_CountRaw": [1.0],
            "GFAP_IntDenTotal": [5.0],
        }
    )

    marker = object.__new__(Antibody)
    marker.name = "GFAP"
    marker.experiment = None
    marker.color = None
    marker.df = pd.DataFrame(
        {
            "AnimalName": ["A1"],
            "Condition": ["A"],
            "GFAP_Volume": [10.0],
            "GFAP_IntDen": [5.0],
            "GFAP_Surface": [2.0],
        }
    )

    behavior = SimpleNamespace(
        df=pd.DataFrame(
            {
                "AnimalName": ["A1"],
                "Condition": ["A"],
                "Period": [24.0],
                "IV": [0.25],
            }
        )
    )

    experiment = SimpleNamespace(
        summary=summary.copy(),
        summaries={"SCN": summary.copy()},
        data={"GFAP": marker},
        name="Exp1",
        filePath=str(tmp_path),
    )

    batch = Batch("Batch1", [experiment], conditions, str(tmp_path))
    batch.data = {"GFAP": marker, "Behaviour": behavior}
    batch.conditions = conditions.conditions
    batch.factor = list(conditions.factor)
    batch.markers = {"GFAP"}
    return batch, summary


def test_export_all_excel_rebuilds_missing_batch_summaries(tmp_path):
    batch, _ = _build_export_batch(tmp_path)
    batch.summaries = {}

    export_dir = tmp_path / "exports"
    batch.export_all_excel(str(export_dir))

    assert (export_dir / "Behavior_Summary.xlsx").exists()
    assert (export_dir / "scn" / "IF_Summary.xlsx").exists()
    assert (export_dir / "scn" / "IF_Extended.xlsx").exists()
    assert (export_dir / "Behavior_Summary_RegexFilters.txt").exists()
    assert (export_dir / "scn" / "IF_Summary_RegexFilters.txt").exists()
    assert (export_dir / "scn" / "IF_Extended_RegexFilters.txt").exists()
    assert "SCN" in batch.summaries


def test_export_if_summary_recovers_legacy_batch_summary_cache(tmp_path):
    batch, summary = _build_export_batch(tmp_path)
    batch.summaries = {}
    batch.__dict__["summary"] = summary.copy()

    export_dir = tmp_path / "summary_only"
    export_dir.mkdir()
    batch.export_IF_summary_excel(str(export_dir))

    assert (export_dir / "IF_Summary.xlsx").exists()
    pd.testing.assert_frame_equal(batch.summaries["SCN"], summary)


def test_export_excel_honors_toggles_and_custom_save_names(tmp_path):
    batch, _ = _build_export_batch(tmp_path)
    export_dir = tmp_path / "custom_exports"
    export_dir.mkdir()

    batch.export_excel(
        str(export_dir),
        behaviour=False,
        if_extended=False,
        if_summary=True,
        if_summary_save_name="CustomSummary",
    )

    assert (export_dir / "scn" / "CustomSummary.xlsx").exists()
    assert (export_dir / "scn" / "CustomSummary_RegexFilters.txt").exists()
    assert not (export_dir / "scn" / "IF_Extended.xlsx").exists()
    assert not (export_dir / "Behavior_Summary.xlsx").exists()


def test_export_if_summary_include_exclude_filters_tabs_and_data_summary(tmp_path):
    batch, _ = _build_export_batch(tmp_path)
    export_dir = tmp_path / "filtered_summary"
    export_dir.mkdir()

    batch.export_IF_summary_excel(
        str(export_dir),
        include=["Count"],
        exclude=["Raw"],
        save_name="FilteredSummary",
    )

    workbook_path = export_dir / "FilteredSummary.xlsx"
    report_path = export_dir / "FilteredSummary_RegexFilters.txt"
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    metric_sheets = [name for name in wb.sheetnames if name not in {"Experimental Conditions", "Data Summary"}]
    assert any("Count per" in name for name in metric_sheets)
    assert not any("Raw Count" in name for name in metric_sheets)
    assert not any("IntDen" in name for name in metric_sheets)

    data_summary_text = str(wb["Data Summary"]["D2"].value)
    assert "Count per" in data_summary_text
    assert "Raw Count" not in data_summary_text
    assert "IntDen" not in data_summary_text

    report_text = report_path.read_text(encoding="utf-8")
    assert "Count" in report_text
    assert "Raw" in report_text


def test_export_behavior_summary_supports_filters_and_custom_name(tmp_path):
    batch, _ = _build_export_batch(tmp_path)
    export_dir = tmp_path / "behavior_only"
    export_dir.mkdir()

    batch.export_behavior_summary_excel(
        str(export_dir),
        include=["Period"],
        save_name="BehaviorOnly",
    )

    workbook_path = export_dir / "BehaviorOnly.xlsx"
    report_path = export_dir / "BehaviorOnly_RegexFilters.txt"
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    metric_sheets = [name for name in wb.sheetnames if name != "Conditions"]
    assert len(metric_sheets) == 1
    assert "period" in metric_sheets[0].lower()

    report_text = report_path.read_text(encoding="utf-8")
    assert "Period" in report_text
    assert "behavior columns" in report_text


def test_export_extended_data_supports_regex_filters_and_report(tmp_path):
    batch, _ = _build_export_batch(tmp_path)
    export_dir = tmp_path / "extended_filtered"
    export_dir.mkdir()

    batch.export_extended_data_excel(
        str(export_dir),
        include=["Volume$"],
        save_name="ExtendedVolumeOnly",
        verbose=False,
    )

    workbook_path = export_dir / "ExtendedVolumeOnly.xlsx"
    report_path = export_dir / "ExtendedVolumeOnly_RegexFilters.txt"
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    sheet_names = [name for name in wb.sheetnames if name != "Conditions"]
    assert len(sheet_names) == 1

    headers = [value for value in next(wb[sheet_names[0]].iter_rows(min_row=1, max_row=1, values_only=True)) if value is not None]
    assert any("Volume" in str(header) for header in headers)
    assert not any("IntDen" in str(header) for header in headers)
    assert not any("SA" in str(header) for header in headers)

    report_text = report_path.read_text(encoding="utf-8")
    assert "Volume$" in report_text
    assert "Experiment 1_GFAP" in report_text


def test_export_all_excel_wrapper_forwards_new_kwargs(tmp_path):
    batch, _ = _build_export_batch(tmp_path)
    export_dir = tmp_path / "wrapper_forwarding"
    export_dir.mkdir()

    batch.export_all_excel(
        str(export_dir),
        behaviour=False,
        if_extended=False,
        if_summary=True,
        if_summary_include=["Count$"],
        if_summary_save_name="CountsOnly",
    )

    assert (export_dir / "scn" / "CountsOnly.xlsx").exists()
    assert not (export_dir / "scn" / "IF_Extended.xlsx").exists()
    assert not (export_dir / "Behavior_Summary.xlsx").exists()


def test_convert_summary_sheet_name_compacts_density_labels_and_descs():
    label, desc = convert_summary_sheet_name("GFAP_VolColoc_DAPI_Count")
    assert label == "GFAP VolColoc DAPI Dens"
    assert "0.1 mm^3" in desc

    label, desc = convert_summary_sheet_name("GFAP_IntDenTotal")
    assert label == "GFAP IntDenDens"
    assert "0.1 mm^3" in desc

    label, desc = convert_summary_sheet_name("GFAP_VolCombo_wDAPI_Count")
    assert label == "GFAP VCmb wDAPI Dens"
    assert "0.1 mm^3" in desc
    assert "marker+" in desc


def test_export_if_summary_compacts_sheet_names_and_reorders_tabs(tmp_path):
    cond_a = condition("A", "A", "#000000", "Genotype", explanation="<>")
    cond_b = condition("B", "B", "#111111", "Genotype", explanation="<>")
    conditions = conditionList([cond_a, cond_b])

    summary = pd.DataFrame(
        {
            "AnimalName": ["A1", "B1"],
            "Condition": ["A", "B"],
            "GFAP_DistToClosest_DAPIMean": [5.0, 6.0],
            "GFAP_burdenScore": [1.2, 1.3],
            "GFAP_VolCombo_wDAPI_Count": [2.0, 2.5],
            "GFAP_VolCombo_wDAPI_CountRaw": [2.0, 2.5],
            "GFAP_VolCombo_wDAPI_IntDenTotal": [7.5, 7.0],
            "GFAP_VolColoc_DAPI_CountRaw": [1.0, 1.1],
            "GFAP_VolColoc_DAPI_Count": [1.5, 1.6],
            "GFAP_IntDenTotal": [9.0, 9.5],
            "GFAP_Count": [3.0, 3.2],
            "GFAP_CountRaw": [3.0, 3.2],
            "GFAP_VolumeTotal": [11.0, 11.5],
        }
    )

    marker = object.__new__(Antibody)
    marker.name = "GFAP"
    marker.experiment = None
    marker.color = None
    marker.df = pd.DataFrame(
        {
            "AnimalName": ["A1", "B1"],
            "Condition": ["A", "B"],
            "GFAP_Volume": [10.0, 11.0],
        }
    )

    behavior = SimpleNamespace(
        df=pd.DataFrame(
            {
                "AnimalName": ["A1", "B1"],
                "Condition": ["A", "B"],
            }
        )
    )

    experiment = SimpleNamespace(
        summary=summary.copy(),
        summaries={"SCN": summary.copy()},
        data={"GFAP": marker},
        name="Exp1",
        filePath=str(tmp_path),
    )

    batch = Batch("Batch1", [experiment], conditions, str(tmp_path))
    batch.data = {"GFAP": marker, "Behaviour": behavior}
    batch.conditions = conditions.conditions
    batch.factor = list(conditions.factor)
    batch.markers = {"GFAP"}

    export_dir = tmp_path / "ordered_summary"
    export_dir.mkdir()
    batch.export_IF_summary_excel(str(export_dir))

    wb = load_workbook(export_dir / "IF_Summary.xlsx", read_only=True, data_only=True)
    metric_sheets = [name for name in wb.sheetnames if name not in {"Experimental Conditions", "Data Summary"}]

    assert "GFAP IntDenDens" in metric_sheets
    assert "GFAP RawVolColoc DAPI" in metric_sheets
    assert "GFAP VCmb wDAPI Dens" in metric_sheets
    assert "GFAP VCmb wDAPI RawCount" in metric_sheets

    assert metric_sheets.index("GFAP IntDenDens") < metric_sheets.index("GFAP Mean Nearest DAPI")
    assert metric_sheets.index("GFAP Mean Nearest DAPI") < metric_sheets.index("GFAP VCmb wDAPI Dens")
